import os

import openpyxl

TestDataValue=""
testDataFile=""
TestDataWk=""
TestDataSh=""
def ReadTestDataFromCrossRef_Category(TestcaseName, fileName):
    path = os.getcwd()
    path= path.replace("\\","/")
    global TestDataValue
    global testDataFile
    global TestDataWk
    global TestDataSh
    try:
        testDataFile = path + "/drop/LennoxPROs/rf-data/" + fileName
        TestDataWk= openpyxl.load_workbook(testDataFile)
        TestDataSh= TestDataWk.active
    except FileNotFoundError:
        testDataFile = path+"/LennoxPROs/rf-data/"+fileName
        TestDataWk = openpyxl.load_workbook(testDataFile)
        TestDataSh = TestDataWk.active
    CreateList=[]
    row= TestDataSh.max_row
    rowValFlag=False
    for i in range (2,row+1):
        CellRef= TestDataSh.cell(row=i, column=1)
        ActTestCaseName= CellRef.value
        if (ActTestCaseName==TestcaseName):
            j=2
            while (True):
                dataRef= TestDataSh.cell(row=i, column=j)
                TestDataValue=dataRef.value
                try:
                    if (not (len(TestDataValue) == 0)):
                        print(TestDataValue)
                        CreateList.append(TestDataValue)
                    j = j + 1
                except TypeError:
                    rowValFlag=True
                    break
        if (rowValFlag==True):
            break
    return CreateList